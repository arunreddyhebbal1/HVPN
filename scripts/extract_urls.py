#!/usr/bin/env python3
"""Extract hyperlink URLs from xlsx files into an adjacent column."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(SCRIPT_DIR))

from sheet_utils import cell_url  # noqa: E402

OUTPUT_SUFFIX = " - with URLs.xlsx"


def find_link_column(ws) -> int:
    for col in range(1, (ws.max_column or 0) + 1):
        header = str(ws.cell(1, col).value or "").strip().lower()
        if "access link" in header or header == "link":
            return col
    return ws.max_column or 1


def output_path_for(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}{OUTPUT_SUFFIX}")


def extract_urls(input_path: Path, output_path: Path | None = None) -> dict[str, object]:
    if output_path is None:
        output_path = output_path_for(input_path)

    wb = load_workbook(input_path)
    ws = wb.active

    link_col = find_link_column(ws)
    url_col = link_col + 1
    if ws.cell(1, url_col).value != "URL":
        ws.cell(1, url_col).value = "URL"

    urls_written = 0
    for row in range(2, (ws.max_row or 1) + 1):
        url = cell_url(ws.cell(row, link_col))
        if url:
            ws.cell(row, url_col).value = url
            urls_written += 1

    wb.save(output_path)

    headers = [
        f"{get_column_letter(col)}={ws.cell(1, col).value!r}"
        for col in range(1, (ws.max_column or 0) + 1)
    ]
    return {
        "input": str(input_path),
        "output": str(output_path),
        "sheet": ws.title,
        "rows": ws.max_row,
        "link_col": get_column_letter(link_col),
        "url_col": get_column_letter(url_col),
        "urls_written": urls_written,
        "headers": headers,
    }


def discover_inputs(root: Path) -> list[Path]:
    files = sorted(root.glob("*.xlsx"))
    return [
        path
        for path in files
        if not path.name.endswith(OUTPUT_SUFFIX)
    ]


def format_result(result: dict[str, object]) -> str:
    return (
        f"input={result['input']}\n"
        f"output={result['output']}\n"
        f"sheet={result['sheet']!r}\n"
        f"rows={result['rows']}\n"
        f"link_col={result['link_col']}\n"
        f"url_col={result['url_col']}\n"
        f"urls_written={result['urls_written']}\n"
        f"headers={result['headers']}"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract hyperlink URLs into adjacent column.")
    parser.add_argument(
        "files",
        nargs="*",
        help="Input .xlsx files (default: all .xlsx in project root)",
    )
    parser.add_argument(
        "--root",
        type=Path,
        default=PROJECT_ROOT,
        help="Directory to scan when no files are passed",
    )
    parser.add_argument(
        "--log",
        type=Path,
        default=PROJECT_ROOT / "data" / "extract_urls_log.txt",
        help="Path for combined log output",
    )
    args = parser.parse_args()

    if args.files:
        inputs = [Path(f) for f in args.files]
    else:
        inputs = discover_inputs(args.root)

    if not inputs:
        raise SystemExit(f"No input .xlsx files found in {args.root}")

    results: list[dict[str, object]] = []
    for input_path in inputs:
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")
        results.append(extract_urls(input_path.resolve()))

    log_blocks = [format_result(result) for result in results]
    log_text = "\n\n".join(log_blocks) + "\n"
    args.log.parent.mkdir(parents=True, exist_ok=True)
    args.log.write_text(log_text, encoding="utf-8")
    print(log_text, end="")


if __name__ == "__main__":
    main()
