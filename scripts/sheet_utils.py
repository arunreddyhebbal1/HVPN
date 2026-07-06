"""Utilities for downloading and parsing nested Google Sheets."""

from __future__ import annotations

import hashlib
import http.client
import random
import re
import socket
import ssl
import time
import urllib.error
import urllib.request
from datetime import date, datetime, time as dt_time, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell

SPREADSHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")
GID_RE = re.compile(r"[#?&]gid=(\d+)")
HYPERLINK_FORMULA_RE = re.compile(
    r'=HYPERLINK\("([^"]+)"(?:,\s*"[^"]*")?\)',
    re.IGNORECASE,
)

LEVEL1_SPREADSHEET_ID = "1DdfIfWczn0VQGW2BebJ__-mwXqyzdVJORM3SfbIbwC0"
LEVEL1_GID = 619238043

FIXED_COLUMNS = [
    "zone",
    "circle",
    "division",
    "substation_title",
    "hierarchy_path",
    "level1_row_index",
    "date",
    "time",
    "level2_sheet_id",
    "level2_gid",
    "level3_sheet_id",
    "level3_gid",
]


def parse_spreadsheet_url(url: str) -> tuple[str, int | None]:
    """Extract spreadsheet ID and optional gid from a Google Sheets URL."""
    if not url:
        raise ValueError("Empty URL")
    match = SPREADSHEET_ID_RE.search(url)
    if not match:
        raise ValueError(f"Could not parse spreadsheet ID from URL: {url}")
    spreadsheet_id = match.group(1)
    gid_match = GID_RE.search(url)
    gid = int(gid_match.group(1)) if gid_match else None
    return spreadsheet_id, gid


def normalize_header(name: str, seen: dict[str, int]) -> str:
    """Trim, collapse whitespace, and deduplicate column names."""
    cleaned = re.sub(r"\s+", " ", str(name).replace("\n", " ").strip())
    if not cleaned:
        cleaned = "unnamed"
    count = seen.get(cleaned, 0)
    seen[cleaned] = count + 1
    if count:
        return f"{cleaned}_{count + 1}"
    return cleaned


def parse_date_value(value: Any) -> date | None:
    """Parse DD.MM.YYYY strings or datetime objects from Level 2."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def format_time_value(value: Any) -> str | None:
    """Normalize hourly timestamps from Level 3."""
    if value is None:
        return None
    if isinstance(value, dt_time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, timedelta) and value.days == 1 and value.seconds == 0:
        return "24:00"
    text = str(value).strip()
    if not text:
        return None
    if text in {"24:00", "24:00:00"}:
        return "24:00"
    return text


def is_hour_row(value: Any) -> bool:
    """Return True when a Level 3 row begins with an hourly time marker."""
    if isinstance(value, (dt_time, datetime)):
        return True
    if isinstance(value, timedelta) and value.days == 1 and value.seconds == 0:
        return True
    if isinstance(value, str) and ":" in value:
        return True
    return False


def cell_url(cell: Cell) -> str | None:
    """Resolve hyperlink target from a worksheet cell."""
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    if isinstance(cell.value, str):
        formula_match = HYPERLINK_FORMULA_RE.match(cell.value.strip())
        if formula_match:
            return formula_match.group(1)
        if cell.value.startswith("http"):
            return cell.value
    return None


def cell_sheet_url(cell: Cell) -> str | None:
    """Resolve a Google Sheets URL from a URL column or hyperlink cell."""
    url = cell_url(cell)
    if url:
        return url
    value = cell.value
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("http"):
            return text
    return None


def find_url_column(headers: list[Any]) -> int:
    """Return 1-based index of the explicit URL column, if present."""
    for idx, header in enumerate(headers, start=1):
        text = str(header or "").strip().lower()
        if text == "url":
            return idx
    return 0


def find_link_column(headers: list[Any]) -> int:
    """Return 1-based index of the access-link column."""
    for idx, header in enumerate(headers, start=1):
        text = str(header or "").strip().lower()
        if "access link" in text or text == "link":
            return idx
    return len(headers)


def find_date_column(headers: list[Any]) -> int:
    """Return 1-based index of the date column."""
    for idx, header in enumerate(headers, start=1):
        text = str(header or "").strip().lower()
        if "date" in text:
            return idx
    return 1


def resolve_url_column(headers: list[Any]) -> int:
    """Prefer the URL column; fall back to Access Link."""
    url_col = find_url_column(headers)
    if url_col:
        return url_col
    return find_link_column(headers)


def is_remote_url(value: str) -> bool:
    text = value.strip().lower()
    return text.startswith("http://") or text.startswith("https://")


def resolve_source_url(cell: Cell) -> str | None:
    """Read a URL from the preferred URL column cell."""
    return cell_sheet_url(cell)


def is_blank_hierarchy_row(zone: str, circle: str, division: str, url: str | None) -> bool:
    """Return True for trailing empty index rows with no hierarchy or link."""
    return not zone and not circle and not division and not url


def fetch_xlsx_bytes(
    source: str,
    *,
    cache_dir: Path | None = None,
    all_sheets: bool = False,
    request_delay: float = 0.0,
) -> bytes:
    """Load workbook bytes from a local path or Google Sheets export URL."""
    if not source or not str(source).strip():
        raise ValueError("Empty source")

    text = str(source).strip()
    if is_remote_url(text):
        spreadsheet_id, gid = parse_spreadsheet_url(text)
        export_gid = None if all_sheets else gid
        return download_xlsx(
            spreadsheet_id,
            export_gid,
            cache_dir=cache_dir,
            request_delay=request_delay,
        )

    path = Path(text)
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")
    return path.read_bytes()


def open_source_workbook(
    source: str,
    *,
    cache_dir: Path | None = None,
    all_sheets: bool = False,
    data_only: bool = True,
    request_delay: float = 0.0,
) -> openpyxl.Workbook:
    """Open a workbook from a local path or remote Google Sheets URL."""
    content = fetch_xlsx_bytes(
        source,
        cache_dir=cache_dir,
        all_sheets=all_sheets,
        request_delay=request_delay,
    )
    return load_workbook_from_bytes(content, data_only=data_only)


def extract_log_worksheet_data(
    worksheet: Any,
) -> tuple[list[str], list[dict[str, Any]]] | None:
    """Parse headers and hourly rows from a Level 3 log tab, or None if empty."""
    max_row = worksheet.max_row or 0
    if max_row < 7:
        return None
    headers = build_l3_headers(worksheet)
    rows = extract_l3_rows(worksheet, headers)
    if not rows:
        return None
    return headers, rows


def worksheet_has_log_data(worksheet: Any) -> bool:
    """Return True when a worksheet looks like a non-empty Level 3 log tab."""
    return extract_log_worksheet_data(worksheet) is not None


def iter_log_worksheet_data(
    workbook: openpyxl.Workbook,
) -> list[tuple[Any, list[str], list[dict[str, Any]]]]:
    """Return Level 3 worksheets with pre-parsed headers and hourly rows."""
    sheets: list[tuple[Any, list[str], list[dict[str, Any]]]] = []
    for worksheet in workbook.worksheets:
        extracted = extract_log_worksheet_data(worksheet)
        if extracted:
            headers, rows = extracted
            sheets.append((worksheet, headers, rows))
    return sheets


def iter_log_worksheets(workbook: openpyxl.Workbook) -> list[Any]:
    """Return Level 3 worksheets that contain hourly log data."""
    return [worksheet for worksheet, _, _ in iter_log_worksheet_data(workbook)]


def substation_name_for_worksheet(worksheet: Any) -> str:
    """Prefer tab title; fall back to the in-sheet substation heading."""
    title = str(worksheet.title or "").strip()
    if title and title.lower() not in {"sheet1", "sheet 1"}:
        return title
    heading = get_substation_title(worksheet)
    return heading or title or "unknown"


def sanitize_path_component(name: str, *, fallback: str = "unknown") -> str:
    """Make a hierarchy or substation label safe for filesystem paths."""
    cleaned = re.sub(r"\s+", " ", str(name or "").strip())
    if not cleaned:
        cleaned = fallback
    cleaned = re.sub(r'[<>:"/\\|?*]', "_", cleaned)
    cleaned = cleaned.strip(" .")
    return cleaned or fallback


def export_url(spreadsheet_id: str, gid: int | None = None) -> str:
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    if gid is not None:
        url += f"&gid={gid}"
    return url


# Transient network failures during long bulk downloads.
DOWNLOAD_RETRY_EXCEPTIONS = (
    urllib.error.URLError,
    urllib.error.HTTPError,
    TimeoutError,
    ConnectionError,
    ConnectionResetError,
    BrokenPipeError,
    http.client.IncompleteRead,
    http.client.RemoteDisconnected,
    socket.timeout,
    ssl.SSLError,
    OSError,
)

NON_RETRYABLE_HTTP_CODES = frozenset({400, 401, 403, 404, 410})


def _unwrap_download_error(exc: BaseException) -> BaseException:
    """Normalize wrapped urllib errors for retry handling."""
    if isinstance(exc, urllib.error.URLError) and exc.reason is not None:
        return exc.reason
    return exc


def _read_response_bytes(response: Any) -> bytes:
    """Read a response body in chunks to reduce IncompleteRead failures."""
    chunks: list[bytes] = []
    while True:
        chunk = response.read(65536)
        if not chunk:
            break
        chunks.append(chunk)
    return b"".join(chunks)


def download_xlsx(
    spreadsheet_id: str,
    gid: int | None = None,
    *,
    cache_dir: Path | None = None,
    retries: int = 6,
    retry_delay: float = 3.0,
    timeout: float = 180.0,
    request_delay: float = 0.0,
) -> bytes:
    """Download spreadsheet as XLSX, optionally using a content-addressed cache."""
    url = export_url(spreadsheet_id, gid)
    cache_key = hashlib.sha256(url.encode("utf-8")).hexdigest()
    cache_path: Path | None = None
    if cache_dir is not None:
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_path = cache_dir / f"{cache_key}.xlsx"
        if cache_path.exists() and cache_path.stat().st_size > 0:
            return cache_path.read_bytes()

    if request_delay > 0:
        time.sleep(request_delay)

    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            request = urllib.request.Request(
                url,
                headers={"User-Agent": "substation-analytics-dashboard/1.0"},
            )
            with urllib.request.urlopen(request, timeout=timeout) as response:
                if isinstance(response, http.client.HTTPResponse):
                    status = response.status
                else:
                    status = getattr(response, "status", 200)
                if status in NON_RETRYABLE_HTTP_CODES:
                    raise RuntimeError(
                        f"HTTP {status} for {url} (not retryable)"
                    )
                content = _read_response_bytes(response)
            if not content:
                raise RuntimeError("Download returned empty response")
            if cache_path is not None:
                cache_path.write_bytes(content)
            return content
        except urllib.error.HTTPError as exc:
            last_error = exc
            if exc.code in NON_RETRYABLE_HTTP_CODES:
                raise RuntimeError(
                    f"HTTP {exc.code} for {url} (not retryable)"
                ) from exc
            if attempt < retries:
                backoff = retry_delay * attempt + random.uniform(0.5, 1.5)
                if exc.code == 429:
                    backoff = retry_delay * (2**attempt) + random.uniform(2.0, 5.0)
                time.sleep(backoff)
        except DOWNLOAD_RETRY_EXCEPTIONS as exc:
            last_error = _unwrap_download_error(exc)  # type: ignore[assignment]
            if attempt < retries:
                backoff = retry_delay * attempt + random.uniform(0.5, 1.5)
                time.sleep(backoff)
    raise RuntimeError(f"Failed to download {url} after {retries} attempts") from last_error


def load_workbook_from_bytes(content: bytes, *, data_only: bool = True) -> openpyxl.Workbook:
    import io

    return openpyxl.load_workbook(io.BytesIO(content), data_only=data_only)


def resolve_sheet(workbook: openpyxl.Workbook, gid: int | None = None) -> Any:
    """Return the active worksheet from an XLSX export.

    Google Sheets XLSX exports include the requested gid as the first (or only) sheet.
    """
    _ = gid
    return workbook.worksheets[0]


def get_substation_title(worksheet: Any) -> str | None:
    value = worksheet.cell(3, 1).value
    if value is None:
        return None
    return str(value).strip()


def build_l3_headers(
    worksheet: Any,
    header_rows: tuple[int, int, int] = (4, 5, 6),
) -> list[str]:
    """Flatten multi-row Level 3 headers into one name per column."""
    headers: list[str] = []
    seen: dict[str, int] = {}
    max_col = worksheet.max_column or 1
    for col in range(1, max_col + 1):
        parts: list[str] = []
        for row in header_rows:
            value = worksheet.cell(row, col).value
            if value is not None and str(value).strip():
                parts.append(str(value).strip())
        label = " | ".join(parts) if parts else f"col_{col}"
        headers.append(normalize_header(label, seen))
    return headers


def extract_l3_rows(worksheet: Any, headers: list[str]) -> list[dict[str, Any]]:
    """Extract hourly data rows from a Level 3 detail sheet."""
    rows: list[dict[str, Any]] = []
    max_row = worksheet.max_row or 0
    for row_idx in range(7, max_row + 1):
        time_value = worksheet.cell(row_idx, 1).value
        if not is_hour_row(time_value):
            continue
        record: dict[str, Any] = {"time": format_time_value(time_value)}
        for col_idx, header in enumerate(headers, start=1):
            if header == "Time":
                continue
            record[header] = worksheet.cell(row_idx, col_idx).value
        rows.append(record)
    return rows


def lookback_bounds(
    *,
    lookback_days: int,
    today: date | None = None,
) -> tuple[date, date]:
    """Return inclusive (lower, upper) date bounds for the lookback window."""
    today = today or date.today()
    lower = today - timedelta(days=lookback_days)
    return lower, today


def within_lookback(
    value: date,
    *,
    lookback_days: int,
    today: date | None = None,
) -> bool:
    """Return True when value is between today-lookback and today."""
    lower, upper = lookback_bounds(lookback_days=lookback_days, today=today)
    return lower <= value <= upper
