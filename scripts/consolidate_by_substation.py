#!/usr/bin/env python3
"""
Consolidate a 3-level Excel / Google Sheets hierarchy into one workbook per substation.

Hierarchy
---------
Level 1 (division master): Zone, Circle, Division, access link, URL
Level 2 (date mapping):    Date, access link, URL
Level 3 (daily logs):      one workbook per date; each tab is a substation log

Output
------
One Excel file per substation under ``data/substations/`` with ~2 years of daily rows:

| Zone | Circle | Division | Date | Substation | time | <log columns...> |

Usage
-----
Quick test (1 division, 2 dates):

    py -3 scripts/consolidate_by_substation.py --division-limit 1 --date-limit 2

Full run (last 730 days, all divisions):

    py -3 scripts/consolidate_by_substation.py

Inspect remote schema:

    py -3 scripts/consolidate_by_substation.py --inspect

Use a local Level-1 workbook instead of the default Google Sheet:

    py -3 scripts/consolidate_by_substation.py --level1 "C:\\path\\to\\master.xlsx"
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, TextIO

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "data" / "substations"
DEFAULT_STAGING_DIR = DEFAULT_OUTPUT_DIR / "_staging"
DEFAULT_ERRORS = PROJECT_ROOT / "data" / "substation_consolidation_errors.csv"
DEFAULT_CACHE = PROJECT_ROOT / "data" / "cache"

from sheet_utils import (  # noqa: E402
    LEVEL1_GID,
    LEVEL1_SPREADSHEET_ID,
    export_url,
    fetch_xlsx_bytes,
    find_date_column,
    is_blank_hierarchy_row,
    iter_log_worksheet_data,
    lookback_bounds,
    open_source_workbook,
    parse_date_value,
    parse_spreadsheet_url,
    resolve_sheet,
    resolve_source_url,
    resolve_url_column,
    sanitize_path_component,
    substation_name_for_worksheet,
)

OUTPUT_FIXED_COLUMNS = ["Zone", "Circle", "Division", "Date", "Substation", "time"]
LOGGER = logging.getLogger("consolidate_by_substation")


@dataclass(frozen=True)
class SubstationKey:
    zone: str
    circle: str
    division: str
    substation: str

    def slug(self) -> str:
        parts = (
            sanitize_path_component(self.zone, fallback="zone"),
            sanitize_path_component(self.circle, fallback="circle"),
            sanitize_path_component(self.division, fallback="division"),
            sanitize_path_component(self.substation, fallback="substation"),
        )
        return "__".join(parts)

    def output_path(self, output_dir: Path) -> Path:
        zone_dir = sanitize_path_component(self.zone, fallback="zone")
        division_dir = sanitize_path_component(self.division, fallback="division")
        filename = f"{sanitize_path_component(self.substation, fallback='substation')}.xlsx"
        return output_dir / zone_dir / division_dir / filename

    def staging_path(self, staging_dir: Path) -> Path:
        digest = hashlib.sha256(self.slug().encode("utf-8")).hexdigest()
        return staging_dir / f"{digest}.jsonl"


@dataclass
class ErrorRecord:
    level1_row_index: int | str
    hierarchy_path: str
    date: str
    url: str
    substation: str
    error: str


@dataclass(frozen=True)
class DateTask:
    level1_row_index: int
    hierarchy: str
    parsed_date: date
    l2_url: str


@dataclass
class DateTaskResult:
    touched_keys: set[SubstationKey] = field(default_factory=set)
    rows_written: int = 0
    substation_tabs_processed: int = 0
    errors: list[ErrorRecord] = field(default_factory=list)
    skipped_empty: bool = False


DEFAULT_RESUME_CHECKPOINT = DEFAULT_STAGING_DIR / "resume_checkpoint.jsonl"


@dataclass
class RunStats:
    level1_rows_processed: int = 0
    level2_dates_found: int = 0
    level2_dates_in_range: int = 0
    level3_workbooks_opened: int = 0
    substation_tabs_processed: int = 0
    rows_written: int = 0
    substations_written: int = 0
    errors: list[ErrorRecord] = field(default_factory=list)


def hierarchy_values(row: list[Any]) -> tuple[str, str, str]:
    zone = str(row[1] or "").strip() if len(row) > 1 else ""
    circle = str(row[2] or "").strip() if len(row) > 2 else ""
    division = str(row[3] or "").strip() if len(row) > 3 else ""
    return zone, circle, division


def hierarchy_path(zone: str, circle: str, division: str) -> str:
    return " > ".join(part for part in (zone, circle, division) if part)


def load_level1_workbook(
    level1_source: str | None,
    *,
    cache_dir: Path,
) -> tuple[Any, Any, list[Any], int]:
    from sheet_utils import download_xlsx, load_workbook_from_bytes

    if level1_source:
        workbook = open_source_workbook(
            level1_source,
            cache_dir=cache_dir,
            data_only=False,
        )
        worksheet = workbook.worksheets[0]
    else:
        content = download_xlsx(LEVEL1_SPREADSHEET_ID, LEVEL1_GID, cache_dir=cache_dir)
        workbook = load_workbook_from_bytes(content, data_only=False)
        worksheet = resolve_sheet(workbook, LEVEL1_GID)

    headers = [
        worksheet.cell(1, col).value
        for col in range(1, (worksheet.max_column or 0) + 1)
    ]
    url_col = resolve_url_column(headers)
    return workbook, worksheet, headers, url_col


def inspect_schema(*, level1_source: str | None, cache_dir: Path) -> dict[str, Any]:
    _, l1_ws, l1_headers, l1_url_col = load_level1_workbook(
        level1_source, cache_dir=cache_dir
    )
    l1_rows = max((l1_ws.max_row or 1) - 1, 0)

    sample_l2 = None
    for row_idx in range(2, (l1_ws.max_row or 1) + 1):
        url = resolve_source_url(l1_ws.cell(row_idx, l1_url_col))
        if url:
            sample_l2 = url
            break
    if not sample_l2:
        raise RuntimeError("Could not find a Level 1 URL for inspection.")

    l2_wb = open_source_workbook(sample_l2, cache_dir=cache_dir, data_only=False)
    l2_ws = l2_wb.worksheets[0]
    l2_headers = [
        l2_ws.cell(1, col).value for col in range(1, (l2_ws.max_column or 0) + 1)
    ]
    date_col = find_date_column(l2_headers)
    l2_url_col = resolve_url_column(l2_headers)

    sample_l3 = None
    for row_idx in range(2, (l2_ws.max_row or 1) + 1):
        url = resolve_source_url(l2_ws.cell(row_idx, l2_url_col))
        if url:
            sample_l3 = (parse_date_value(l2_ws.cell(row_idx, date_col).value), url)
            break
    if not sample_l3:
        raise RuntimeError("Could not find a Level 2 URL for inspection.")

    l3_wb = open_source_workbook(
        sample_l3[1],
        cache_dir=cache_dir,
        all_sheets=True,
        data_only=True,
    )
    log_data = iter_log_worksheet_data(l3_wb)
    if not log_data:
        raise RuntimeError("Level 3 workbook has no non-empty log tabs.")

    first_tab, l3_headers, l3_rows_list = log_data[0]
    l3_rows = len(l3_rows_list)

    return {
        "level1": {
            "source": level1_source or export_url(LEVEL1_SPREADSHEET_ID, LEVEL1_GID),
            "columns": l1_headers,
            "url_column_index": l1_url_col,
            "row_count": l1_rows,
        },
        "level2": {
            "sample_url": sample_l2,
            "columns": l2_headers,
            "url_column_index": l2_url_col,
            "row_count": max((l2_ws.max_row or 1) - 1, 0),
            "date_format": "DD.MM.YYYY",
        },
        "level3": {
            "sample_url": sample_l3[1],
            "tab_count": len(l3_wb.worksheets),
            "log_tab_count": len(log_data),
            "log_tab_names": [ws.title for ws, _, _ in log_data[:20]],
            "sample_tab": first_tab.title,
            "sample_substation": substation_name_for_worksheet(first_tab),
            "column_count": len(l3_headers),
            "hourly_rows_per_date": l3_rows,
        },
    }


def append_staged_row(staging_path: Path, row: dict[str, Any]) -> None:
    staging_path.parent.mkdir(parents=True, exist_ok=True)
    with staging_path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(row, default=str) + "\n")


def load_processed_dates_from_staging(staging_dir: Path) -> set[tuple[str, str]]:
    processed: set[tuple[str, str]] = set()
    if not staging_dir.is_dir():
        return processed
    for staging_path in staging_dir.glob("*.jsonl"):
        if staging_path.name == "resume_checkpoint.jsonl":
            continue
        with staging_path.open("r", encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if not line:
                    continue
                row = json.loads(line)
                path = hierarchy_path(
                    str(row.get("Zone", "")),
                    str(row.get("Circle", "")),
                    str(row.get("Division", "")),
                )
                date_value = str(row.get("Date", ""))
                if path and date_value:
                    processed.add((path, date_value))
    return processed


def load_resume_checkpoint(checkpoint_path: Path) -> set[str]:
    if not checkpoint_path.exists():
        return set()
    urls: set[str] = set()
    with checkpoint_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            record = json.loads(line)
            url = record.get("l2_url")
            if url:
                urls.add(str(url))
    return urls


def append_resume_checkpoint(
    checkpoint_path: Path,
    *,
    l2_url: str,
    hierarchy: str,
    parsed_date: date,
) -> None:
    checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
    record = {
        "l2_url": l2_url,
        "hierarchy_path": hierarchy,
        "date": parsed_date.isoformat(),
    }
    with checkpoint_path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record) + "\n")


def process_l3_date(
    task: DateTask,
    *,
    zone: str,
    circle: str,
    division: str,
    path: str,
    level1_row_index: int,
    cache_dir: Path,
    staging_dir: Path,
    request_delay: float,
    staging_lock: threading.Lock,
    resume_checkpoint_path: Path,
) -> DateTaskResult:
    result = DateTaskResult()
    try:
        l3_wb = open_source_workbook(
            task.l2_url,
            cache_dir=cache_dir,
            all_sheets=True,
            data_only=True,
            request_delay=request_delay,
        )
        log_data = iter_log_worksheet_data(l3_wb)
        if not log_data:
            result.errors.append(
                ErrorRecord(
                    level1_row_index,
                    path,
                    task.parsed_date.isoformat(),
                    task.l2_url,
                    "",
                    "No non-empty log tabs in Level 3 workbook",
                )
            )
            return result
    except Exception as exc:
        result.errors.append(
            ErrorRecord(
                level1_row_index,
                path,
                task.parsed_date.isoformat(),
                task.l2_url,
                "",
                str(exc),
            )
        )
        return result

    with staging_lock:
        for worksheet, _headers, detail_rows in log_data:
            substation = substation_name_for_worksheet(worksheet)
            key = SubstationKey(zone, circle, division, substation)
            result.touched_keys.add(key)
            result.substation_tabs_processed += 1

            if not detail_rows:
                continue

            staging_path = key.staging_path(staging_dir)
            for detail in detail_rows:
                output_row: dict[str, Any] = {
                    "Zone": zone,
                    "Circle": circle,
                    "Division": division,
                    "Date": task.parsed_date.isoformat(),
                    "Substation": substation,
                    "time": detail.get("time", ""),
                }
                for col_name, value in detail.items():
                    if col_name == "time":
                        continue
                    output_row[col_name] = value
                append_staged_row(staging_path, output_row)
                result.rows_written += 1

        append_resume_checkpoint(
            resume_checkpoint_path,
            l2_url=task.l2_url,
            hierarchy=path,
            parsed_date=task.parsed_date,
        )

    return result


def log_error(
    error_writer: csv.DictWriter,
    err_file: TextIO,
    stats: RunStats,
    *,
    level1_row_index: int | str,
    hierarchy: str,
    date_value: str,
    url: str,
    substation: str,
    error: str,
) -> None:
    record = ErrorRecord(
        level1_row_index, hierarchy, date_value, url, substation, error
    )
    stats.errors.append(record)
    error_writer.writerow(
        {
            "level1_row_index": level1_row_index,
            "hierarchy_path": hierarchy,
            "date": date_value,
            "url": url,
            "substation": substation,
            "error": error,
        }
    )
    err_file.flush()


def consolidate(
    *,
    output_dir: Path,
    staging_dir: Path,
    errors_csv: Path,
    cache_dir: Path,
    level1_source: str | None,
    lookback_days: int,
    division_limit: int | None,
    date_limit: int | None,
    request_delay: float,
    skip_excel_write: bool,
    workers: int,
    resume: bool,
    division_checkpoint: bool,
    resume_checkpoint_path: Path,
) -> RunStats:
    stats = RunStats()
    lower_bound, upper_bound = lookback_bounds(lookback_days=lookback_days)
    touched_keys: set[SubstationKey] = set()
    processed_urls = load_resume_checkpoint(resume_checkpoint_path) if resume else set()
    processed_dates = (
        load_processed_dates_from_staging(staging_dir) if resume else set()
    )
    staging_lock = threading.Lock()

    if resume and (processed_urls or processed_dates):
        LOGGER.info(
            "Resume: skipping %s checkpoint URL(s), %s staged date(s)",
            len(processed_urls),
            len(processed_dates),
        )

    _, l1_ws, l1_headers, l1_url_col = load_level1_workbook(
        level1_source, cache_dir=cache_dir
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    staging_dir.mkdir(parents=True, exist_ok=True)
    errors_csv.parent.mkdir(parents=True, exist_ok=True)

    with errors_csv.open("w", newline="", encoding="utf-8") as err_file:
        error_writer = csv.DictWriter(
            err_file,
            fieldnames=[
                "level1_row_index",
                "hierarchy_path",
                "date",
                "url",
                "substation",
                "error",
            ],
        )
        error_writer.writeheader()

        divisions_processed = 0
        for row_idx in range(2, (l1_ws.max_row or 1) + 1):
            if division_limit is not None and divisions_processed >= division_limit:
                break

            row_values = [
                l1_ws.cell(row_idx, col).value
                for col in range(1, (l1_ws.max_column or 0) + 1)
            ]
            zone, circle, division = hierarchy_values(row_values)
            path = hierarchy_path(zone, circle, division)
            l1_url = resolve_source_url(l1_ws.cell(row_idx, l1_url_col))
            if is_blank_hierarchy_row(zone, circle, division, l1_url):
                continue
            if not l1_url:
                log_error(
                    error_writer,
                    err_file,
                    stats,
                    level1_row_index=row_idx,
                    hierarchy=path,
                    date_value="",
                    url="",
                    substation="",
                    error="Missing Level 1 URL",
                )
                continue

            stats.level1_rows_processed += 1
            divisions_processed += 1
            LOGGER.info("[L1 %s] %s", divisions_processed, path)

            try:
                l2_wb = open_source_workbook(
                    l1_url,
                    cache_dir=cache_dir,
                    data_only=False,
                    request_delay=request_delay,
                )
                l2_ws = l2_wb.worksheets[0]
                l2_headers = [
                    l2_ws.cell(1, col).value
                    for col in range(1, (l2_ws.max_column or 0) + 1)
                ]
                date_col = find_date_column(l2_headers)
                l2_url_col = resolve_url_column(l2_headers)
            except Exception as exc:
                log_error(
                    error_writer,
                    err_file,
                    stats,
                    level1_row_index=row_idx,
                    hierarchy=path,
                    date_value="",
                    url=l1_url,
                    substation="",
                    error=str(exc),
                )
                continue

            division_keys: set[SubstationKey] = set()
            date_tasks: list[DateTask] = []
            dates_collected = 0

            for l2_row_idx in range(l2_ws.max_row or 1, 1, -1):
                if date_limit is not None and dates_collected >= date_limit:
                    break

                raw_date = l2_ws.cell(l2_row_idx, date_col).value
                parsed_date = parse_date_value(raw_date)
                if parsed_date is None:
                    continue

                stats.level2_dates_found += 1
                if parsed_date < lower_bound:
                    break
                if parsed_date > upper_bound:
                    continue

                stats.level2_dates_in_range += 1
                l2_url = resolve_source_url(l2_ws.cell(l2_row_idx, l2_url_col))
                if not l2_url:
                    log_error(
                        error_writer,
                        err_file,
                        stats,
                        level1_row_index=row_idx,
                        hierarchy=path,
                        date_value=parsed_date.isoformat(),
                        url="",
                        substation="",
                        error="Missing Level 2 URL",
                    )
                    continue

                if resume and l2_url in processed_urls:
                    continue
                if resume and (path, parsed_date.isoformat()) in processed_dates:
                    continue

                dates_collected += 1
                date_tasks.append(
                    DateTask(
                        level1_row_index=row_idx,
                        hierarchy=path,
                        parsed_date=parsed_date,
                        l2_url=l2_url,
                    )
                )

            if not date_tasks:
                LOGGER.info("  no dates to process for this division")
                continue

            dates_done = 0
            with ThreadPoolExecutor(max_workers=workers) as executor:
                futures = {
                    executor.submit(
                        process_l3_date,
                        task,
                        zone=zone,
                        circle=circle,
                        division=division,
                        path=path,
                        level1_row_index=row_idx,
                        cache_dir=cache_dir,
                        staging_dir=staging_dir,
                        request_delay=request_delay,
                        staging_lock=staging_lock,
                        resume_checkpoint_path=resume_checkpoint_path,
                    ): task
                    for task in date_tasks
                }
                for future in as_completed(futures):
                    task = futures[future]
                    dates_done += 1
                    try:
                        result = future.result()
                    except Exception as exc:
                        log_error(
                            error_writer,
                            err_file,
                            stats,
                            level1_row_index=row_idx,
                            hierarchy=path,
                            date_value=task.parsed_date.isoformat(),
                            url=task.l2_url,
                            substation="",
                            error=str(exc),
                        )
                        continue

                    division_keys.update(result.touched_keys)
                    touched_keys.update(result.touched_keys)
                    stats.level3_workbooks_opened += 1
                    stats.substation_tabs_processed += result.substation_tabs_processed
                    stats.rows_written += result.rows_written
                    processed_urls.add(task.l2_url)

                    for err in result.errors:
                        log_error(
                            error_writer,
                            err_file,
                            stats,
                            level1_row_index=err.level1_row_index,
                            hierarchy=err.hierarchy_path,
                            date_value=err.date,
                            url=err.url,
                            substation=err.substation,
                            error=err.error,
                        )

                    if dates_done % 25 == 0:
                        LOGGER.info(
                            "  dates %s/%s | rows %s | substations %s | errors %s",
                            dates_done,
                            len(date_tasks),
                            stats.rows_written,
                            len(touched_keys),
                            len(stats.errors),
                        )

            if division_checkpoint and not skip_excel_write and division_keys:
                written = write_substation_workbooks(
                    division_keys, staging_dir=staging_dir, output_dir=output_dir
                )
                stats.substations_written += written
                LOGGER.info(
                    "  division checkpoint: wrote %s substation workbook(s)",
                    written,
                )

    if not skip_excel_write and not division_checkpoint:
        stats.substations_written = write_substation_workbooks(
            touched_keys, staging_dir=staging_dir, output_dir=output_dir
        )

    return stats


def read_staged_rows(staging_path: Path) -> list[dict[str, Any]]:
    if not staging_path.exists():
        return []
    rows: list[dict[str, Any]] = []
    with staging_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def union_fieldnames(rows: list[dict[str, Any]]) -> list[str]:
    dynamic: list[str] = []
    seen = set(OUTPUT_FIXED_COLUMNS)
    for row in rows:
        for key in row:
            if key in seen:
                continue
            seen.add(key)
            dynamic.append(key)
    return OUTPUT_FIXED_COLUMNS + dynamic


def write_substation_workbooks(
    keys: set[SubstationKey],
    *,
    staging_dir: Path,
    output_dir: Path,
) -> int:
    import openpyxl

    written = 0
    for key in sorted(keys, key=lambda item: item.slug()):
        staging_path = key.staging_path(staging_dir)
        rows = read_staged_rows(staging_path)
        if not rows:
            continue

        rows.sort(key=lambda row: (row.get("Date", ""), row.get("time", "")))
        fieldnames = union_fieldnames(rows)
        output_path = key.output_path(output_dir)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "data"
        worksheet.append(fieldnames)
        for row in rows:
            worksheet.append([row.get(name, "") for name in fieldnames])
        workbook.save(output_path)
        written += 1
        LOGGER.info("Wrote %s (%s rows)", output_path, len(rows))

    return written


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory for per-substation Excel files",
    )
    parser.add_argument(
        "--staging-dir",
        type=Path,
        default=DEFAULT_STAGING_DIR,
        help="Temporary JSONL staging directory",
    )
    parser.add_argument(
        "--errors",
        type=Path,
        default=DEFAULT_ERRORS,
        help="Error log CSV output path",
    )
    parser.add_argument(
        "--cache-dir",
        type=Path,
        default=DEFAULT_CACHE,
        help="Directory for downloaded workbook cache files",
    )
    parser.add_argument(
        "--level1",
        type=str,
        default=None,
        help="Optional local Level-1 workbook path (defaults to Google Sheet master)",
    )
    parser.add_argument(
        "--lookback-days",
        type=int,
        default=730,
        help="Only include dates within this many days of today",
    )
    parser.add_argument(
        "--division-limit",
        type=int,
        default=None,
        help="Process only the first N Level 1 divisions (for testing)",
    )
    parser.add_argument(
        "--date-limit",
        type=int,
        default=None,
        help="Process only the first N in-range dates per division (for testing)",
    )
    parser.add_argument(
        "--request-delay",
        type=float,
        default=0.15,
        help="Seconds to wait before each uncached remote workbook download",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=4,
        help="Parallel workers for Level 3 downloads within each division",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Skip Level 2 URLs already recorded in the resume checkpoint",
    )
    parser.add_argument(
        "--resume-checkpoint",
        type=Path,
        default=DEFAULT_RESUME_CHECKPOINT,
        help="JSONL file tracking successfully processed Level 2 URLs",
    )
    parser.add_argument(
        "--division-checkpoint",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Write Excel workbooks after each division completes (default: on)",
    )
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="Print discovered schema JSON and exit",
    )
    parser.add_argument(
        "--skip-excel-write",
        action="store_true",
        help="Only stage JSONL rows; skip final Excel export",
    )
    parser.add_argument(
        "--from-staging",
        action="store_true",
        help="Rebuild Excel files from existing staging JSONL without re-downloading",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
    )
    return parser


def rebuild_from_staging(staging_dir: Path, output_dir: Path) -> int:
    keys: set[SubstationKey] = set()
    for staging_path in staging_dir.glob("*.jsonl"):
        rows = read_staged_rows(staging_path)
        if not rows:
            continue
        sample = rows[0]
        keys.add(
            SubstationKey(
                str(sample.get("Zone", "")),
                str(sample.get("Circle", "")),
                str(sample.get("Division", "")),
                str(sample.get("Substation", "")),
            )
        )
    return write_substation_workbooks(keys, staging_dir=staging_dir, output_dir=output_dir)


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(message)s",
    )

    if args.inspect:
        schema = inspect_schema(level1_source=args.level1, cache_dir=args.cache_dir)
        print(json.dumps(schema, indent=2))
        return 0

    if args.from_staging:
        count = rebuild_from_staging(args.staging_dir, args.output_dir)
        print(f"Rebuilt {count} substation workbook(s) in {args.output_dir}")
        return 0 if count > 0 else 1

    LOGGER.info(
        "Starting per-substation consolidation (lookback=%s days, division_limit=%s, date_limit=%s, workers=%s, resume=%s)",
        args.lookback_days,
        args.division_limit,
        args.date_limit,
        args.workers,
        args.resume,
    )
    stats = consolidate(
        output_dir=args.output_dir,
        staging_dir=args.staging_dir,
        errors_csv=args.errors,
        cache_dir=args.cache_dir,
        level1_source=args.level1,
        lookback_days=args.lookback_days,
        division_limit=args.division_limit,
        date_limit=args.date_limit,
        request_delay=args.request_delay,
        skip_excel_write=args.skip_excel_write,
        workers=max(1, args.workers),
        resume=args.resume,
        division_checkpoint=args.division_checkpoint,
        resume_checkpoint_path=args.resume_checkpoint,
    )

    print("Run complete:")
    print(f"  Level 1 rows processed:   {stats.level1_rows_processed}")
    print(f"  Level 2 dates found:      {stats.level2_dates_found}")
    print(f"  Level 2 dates in range:   {stats.level2_dates_in_range}")
    print(f"  Level 3 workbooks opened: {stats.level3_workbooks_opened}")
    print(f"  Substation tabs processed:{stats.substation_tabs_processed}")
    print(f"  Rows written:             {stats.rows_written}")
    print(f"  Substation files written:{stats.substations_written}")
    print(f"  Errors logged:            {len(stats.errors)}")
    print(f"  Output directory:         {args.output_dir}")
    print(f"  Errors:                   {args.errors}")
    return 0 if stats.rows_written > 0 else 1


if __name__ == "__main__":
    sys.exit(main())
